"""
Outputs: charts (PNG), raw data ready for Excel (CSV + XLSX), summaries.

Note on the CSV files: the text is English, but the numbers are written
the Czech way - semicolon separator, decimal comma, UTF-8 BOM - so they
open correctly on a Czech Windows without the import wizard.
"""
from __future__ import annotations

import math
import os
import re
import sys
from pathlib import Path

import matplotlib
import numpy as np
import pandas as pd

# Always render through Agg unless the caller explicitly asked for an
# on-screen window (SF_SHOW_CHARTS=1, set by the --show path).
#
# The GUI runs the analysis on a worker thread, and every interactive
# backend - TkAgg on Windows, MacOSX on a Mac - has to be driven from the
# main thread. Creating a figure off it produces "Starting a Matplotlib GUI
# outside of the main thread will likely fail" and then, depending on the
# machine, no PNG at all. Writing files never needs a GUI backend, so the
# default is the one that always works and the window is opt-in.
if os.environ.get("SF_SHOW_CHARTS") != "1":
    matplotlib.use("Agg")
elif not (os.environ.get("DISPLAY") or sys.platform.startswith(("win", "darwin"))):
    matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
from matplotlib.ticker import AutoMinorLocator  # noqa: E402

from sf_analysis import PCT_STEPS, Result, summary_rows  # noqa: E402
from sf_protocol import Burn  # noqa: E402

# --- palette -----------------------------------------------------------
# Categorical slots are assigned in a fixed order and never cycled.
C_THRUST = "#2a78d6"   # slot 1, blue - measured thrust
C_BASE = "#eb6834"     # slot 2, orange - mass-loss correction
C_IMPULSE = "#1baf7a"  # slot 3, aqua - cumulative impulse
C_CRIT = "#d03b3b"     # status critical - pyro / peak
C_MUTED = "#898781"
C_GRID = "#e1e0d9"
C_AXIS = "#c3c2b7"
C_INK = "#0b0b0b"
C_INK2 = "#52514e"
C_SURFACE = "#fcfcfb"

SERIES = [C_THRUST, C_BASE, C_IMPULSE, "#eda100", "#e87ba4", "#008300", "#4a3aa7", "#e34948"]


def _style(ax, xlabel: str, ylabel: str, title: str | None = None) -> None:
    ax.set_facecolor(C_SURFACE)
    ax.grid(True, color=C_GRID, linewidth=0.8, zorder=0)
    ax.set_axisbelow(True)
    for side in ("top", "right"):
        ax.spines[side].set_visible(False)
    for side in ("left", "bottom"):
        ax.spines[side].set_color(C_AXIS)
        ax.spines[side].set_linewidth(1.0)
    ax.tick_params(colors=C_MUTED, labelsize=9, length=3)
    ax.xaxis.set_minor_locator(AutoMinorLocator(2))
    ax.set_xlabel(xlabel, color=C_INK2, fontsize=10)
    ax.set_ylabel(ylabel, color=C_INK2, fontsize=10)
    if title:
        ax.set_title(title, color=C_INK, fontsize=12, fontweight="600", loc="left", pad=10)


def _ok(v) -> bool:
    return v is not None and isinstance(v, float) and not math.isnan(v)


def _info_line(burn: Burn, r: Result) -> str:
    """Propellant mass, Isp and the calibration factor used for this burn -
    the three numbers people ask about first when comparing charts."""
    mass = f"{r.prop_mass_kg * 1000:.1f} g" if r.prop_mass_kg > 0 else "n/a"
    isp = f"{r.isp:.1f} s" if _ok(r.isp) else "n/a"
    return (f"Propellant {mass}   .   Isp {isp}   .   "
            f"Cal {burn.cal_counts_per_n:.1f} counts/N")


# ---------------------------------------------------------------------
#  The chart sheet
# ---------------------------------------------------------------------
def plot_burn(burn: Burn, r: Result, outdir: Path, show: bool = False) -> Path:
    df = burn.df
    fig = plt.figure(figsize=(13, 11), facecolor=C_SURFACE)
    gs = fig.add_gridspec(3, 2, height_ratios=[1.5, 1, 1], hspace=0.42, wspace=0.22)

    fig.suptitle(f"Static fire #{r.burn_id}   .   {r.motor_designation}   .   "
                 f"{r.total_impulse:.1f} N.s",
                 color=C_INK, fontsize=17, fontweight="700", x=0.045, ha="left", y=0.985)
    fig.text(0.045, 0.958, _info_line(burn, r), color=C_INK2, fontsize=11,
             ha="left", va="top")

    # The record is normally much longer than the burn, so the charts are
    # cropped around the interesting part - otherwise the curve would be a
    # thin spike at the edge of the picture.
    t_min = float(df["t"].min())
    t_max = float(df["t"].max())
    t_view = min(t_max, max(r.burn_time * 1.6, r.burn_time + 2.0)) if r.burn_time > 0 else t_max
    clipped = t_view < t_max - 0.5

    # ---- 1) the whole thrust curve ------------------------------------
    ax = fig.add_subplot(gs[0, :])
    _style(ax, "Time from fire command [s]", "Thrust [N]",
           "Thrust curve" + (f"  (record continues to {t_max:.0f} s)" if clipped else ""))

    if not r.curve.empty:
        span = df.loc[df["t"] <= t_view, "thrust"]
        lo, hi = float(span.min()), float(span.max())
        pad = max(0.05 * (hi - lo), 0.5)

        pyro = df[df["pyro"]]
        if len(pyro):
            ax.axvspan(float(pyro["t"].iloc[0]), float(pyro["t"].iloc[-1]),
                       color=C_CRIT, alpha=0.10, zorder=1,
                       label=f"pyro energised ({float(pyro['t'].iloc[-1]):.2f} s)")

        if r.action_time > 0 and _ok(r.pct_times.get(5)):
            ax.axvspan(r.pct_times[5], r.burn_time, color=C_THRUST, alpha=0.07, zorder=1,
                       label=f"action time ({r.action_time:.2f} s)")

        ax.plot(df["t"], df["thrust"], color=C_THRUST, linewidth=2.0, zorder=4,
                label="measured thrust")
        ax.plot(r.curve["t"], r.curve["baseline"], color=C_BASE, linewidth=2.0,
                linestyle=(0, (4, 3)), zorder=5, label="mass-loss correction")

        ax.axvline(0, color=C_MUTED, linewidth=1.2, linestyle=(0, (2, 3)), zorder=3)
        ax.annotate("T0", xy=(0, lo - pad), xytext=(4, 6), textcoords="offset points",
                    color=C_MUTED, fontsize=9, va="bottom")

        if _ok(r.t_peak):
            ax.plot([r.t_peak], [r.baseline_n + r.peak_thrust], "o", markersize=9,
                    color=C_CRIT, markeredgecolor=C_SURFACE, markeredgewidth=2, zorder=6)
            ax.annotate(f"peak {r.peak_thrust:.1f} N  @ {r.t_peak:.2f} s",
                        xy=(r.t_peak, r.baseline_n + r.peak_thrust),
                        xytext=(12, 2), textcoords="offset points",
                        color=C_INK, fontsize=10, fontweight="600", va="center")
        if _ok(r.t_first_motion):
            ax.plot([r.t_first_motion], [r.baseline_n + r.detection_threshold_n], "o",
                    markersize=8, color=C_IMPULSE, markeredgecolor=C_SURFACE,
                    markeredgewidth=2, zorder=6)
            ax.annotate(f"first motion  {r.t_first_motion*1000:.0f} ms",
                        xy=(r.t_first_motion, r.baseline_n + r.detection_threshold_n),
                        xytext=(10, 12), textcoords="offset points",
                        color=C_INK2, fontsize=9)

        ax.set_xlim(t_min, t_view)
        ax.set_ylim(lo - pad, hi + pad * 2.2)
        leg = ax.legend(loc="upper right", frameon=False, fontsize=9)
        for txt in leg.get_texts():
            txt.set_color(C_INK2)

    # ---- 2) ignition close-up ------------------------------------------
    ax = fig.add_subplot(gs[1, 0])
    _style(ax, "Time from fire command [s]", "Thrust above rest [N]", "Ignition close-up")
    zoom_hi = max(0.6, (r.t_peak * 1.7) if _ok(r.t_peak) else 0.6)
    z = df[(df["t"] >= -0.25) & (df["t"] <= zoom_hi)]
    if len(z):
        ax.plot(z["t"], z["thrust"] - r.baseline_n, color=C_THRUST, linewidth=2.0,
                marker="o", markersize=3.2, markeredgewidth=0, zorder=4)
        ax.axhline(r.detection_threshold_n, color=C_MUTED, linewidth=1.0,
                   linestyle=(0, (2, 3)), zorder=3)
        ax.axvline(0, color=C_CRIT, linewidth=1.4, zorder=3)
        ax.annotate("command", xy=(0, ax.get_ylim()[0]), xytext=(4, 6),
                    textcoords="offset points", color=C_CRIT, fontsize=9, va="bottom")
        for pct in (10, 50, 90):
            tv = r.pct_times.get(pct)
            if _ok(tv) and tv <= zoom_hi:
                ax.axvline(tv, color=C_MUTED, linewidth=0.9, linestyle=(0, (1, 3)), zorder=2)
        if _ok(r.t_first_motion):
            ax.axvline(r.t_first_motion, color=C_IMPULSE, linewidth=1.6, zorder=3)

        # The igniter blip, if one was set aside. Drawn so the decision can
        # be checked by eye rather than taken on trust.
        if _ok(r.t_igniter_spike):
            ax.plot([r.t_igniter_spike], [r.igniter_spike_n], "x", markersize=9,
                    markeredgewidth=2.0, color=C_MUTED, zorder=6)
            ax.annotate("igniter", xy=(r.t_igniter_spike, r.igniter_spike_n),
                        xytext=(6, 4), textcoords="offset points",
                        color=C_MUTED, fontsize=8)

        # The thresholds sit tens of milliseconds apart, so labels next to
        # each line would overlap; they go into one block instead.
        lines = []
        if _ok(r.t_igniter_spike):
            lines.append(f"igniter blip  {r.t_igniter_spike*1000:7.0f} ms")
        if _ok(r.t_first_motion):
            lines.append(f"first motion  {r.t_first_motion*1000:7.0f} ms")
        for pct in (10, 50, 90):
            tv = r.pct_times.get(pct)
            if _ok(tv):
                lines.append(f"{pct:>3} % of peak {tv*1000:7.0f} ms")
        if _ok(r.rise_10_90):
            lines.append(f"rise 10-90 %  {r.rise_10_90*1000:7.0f} ms")
        # Drop the block into whichever corner the trace actually leaves
        # free. Guessing from the peak position alone put it straight on top
        # of the curve as soon as the motor was already running at the left
        # edge (an early ignition fills the whole panel), so count how many
        # samples fall in each corner and take the emptiest.
        zt = z["t"].to_numpy(dtype=float)
        zn = (z["thrust"] - r.baseline_n).to_numpy(dtype=float)
        # np.ptp(), not ndarray.ptp() - the method was removed in NumPy 2.
        xs = (zt - zt.min()) / max(float(np.ptp(zt)), 1e-9)
        ys = (zn - zn.min()) / max(float(np.ptp(zn)), 1e-9)
        corners = {                       # (x anchor, ha, y anchor, va)
            (0.025, "left", 0.97, "top"):     ((xs < 0.45) & (ys > 0.55)),
            (0.975, "right", 0.97, "top"):    ((xs > 0.55) & (ys > 0.55)),
            (0.025, "left", 0.03, "bottom"):  ((xs < 0.45) & (ys < 0.45)),
            (0.975, "right", 0.03, "bottom"): ((xs > 0.55) & (ys < 0.45)),
        }
        bx, ha, by, va = min(corners, key=lambda k: int(corners[k].sum()))
        ax.text(bx, by, "\n".join(lines), transform=ax.transAxes,
                ha=ha, va=va, fontsize=9, color=C_INK2, family="monospace",
                bbox=dict(boxstyle="round,pad=0.5", facecolor=C_SURFACE,
                          edgecolor=C_GRID))

    # ---- 3) cumulative impulse -----------------------------------------
    ax = fig.add_subplot(gs[1, 1])
    _style(ax, "Time from fire command [s]", "Impulse [N.s]", "Cumulative impulse")
    if not r.curve.empty:
        c = r.curve
        net = np.maximum(c["thrust"].to_numpy() - c["baseline"].to_numpy(), 0.0)
        t = c["t"].to_numpy()
        cum = np.concatenate([[0.0], np.cumsum(0.5 * (net[1:] + net[:-1]) * np.diff(t))])
        ax.plot(t, cum, color=C_IMPULSE, linewidth=2.0, zorder=4)
        ax.fill_between(t, 0, cum, color=C_IMPULSE, alpha=0.10, zorder=2)
        # Starts where the integration does, which is before T0 when the
        # motor lit early - clipping at 0 there would draw the curve already
        # part way up and make it look like impulse appeared from nowhere.
        ax.set_xlim(min(0.0, float(t[0])), t_view)
        if _ok(r.t_half_impulse):
            ax.plot([r.t_half_impulse], [np.interp(r.t_half_impulse, t, cum)], "o",
                    markersize=8, color=C_INK, markeredgecolor=C_SURFACE,
                    markeredgewidth=2, zorder=5)
            ax.annotate(f"50 % of impulse\n{r.t_half_impulse:.2f} s",
                        xy=(r.t_half_impulse, np.interp(r.t_half_impulse, t, cum)),
                        xytext=(8, -22), textcoords="offset points",
                        color=C_INK2, fontsize=9)
        ax.annotate(f"{r.total_impulse:.1f} N.s total", xy=(t_view, cum[-1]),
                    xytext=(-6, -14), textcoords="offset points", ha="right",
                    color=C_INK, fontsize=10, fontweight="600")

    # ---- 4) rise thresholds --------------------------------------------
    ax = fig.add_subplot(gs[2, 0])
    _style(ax, "Time from fire command [ms]", "Fraction of peak thrust [%]", "Thrust build-up")
    pcts = [p for p in PCT_STEPS if _ok(r.pct_times.get(p))]
    if pcts:
        xs = [r.pct_times[p] * 1000.0 for p in pcts]
        ax.plot(xs, pcts, color=C_THRUST, linewidth=2.0, marker="o", markersize=7,
                markeredgecolor=C_SURFACE, markeredgewidth=2, zorder=4)
        # label only some points, otherwise they run into each other
        for x, p in zip(xs, pcts):
            if p in (5, 25, 75, 100):
                ax.annotate(f"{x:.0f} ms", xy=(x, p), xytext=(8, -3),
                            textcoords="offset points", color=C_INK2, fontsize=9)
        ax.set_ylim(0, 108)
        ax.set_xlim(min(xs) - 0.14 * (max(xs) - min(xs) + 1),
                    max(xs) + 0.30 * (max(xs) - min(xs) + 1))

    # ---- 5) data quality ------------------------------------------------
    ax = fig.add_subplot(gs[2, 1])
    _style(ax, "Time from fire command [s]", "Interval between samples [ms]", "Record quality")
    if len(df) > 2:
        dt = np.diff(df["t"].to_numpy()) * 1000.0
        ax.plot(df["t"].to_numpy()[1:], dt, color=C_MUTED, linewidth=1.0, zorder=3)
        nominal = float(np.median(dt))
        ax.axhline(nominal, color=C_THRUST, linewidth=1.6, zorder=4)
        ax.annotate(f"median {nominal:.1f} ms  ({1000/nominal:.0f} Hz)",
                    xy=(df["t"].iloc[-1], nominal), xytext=(-6, 6),
                    textcoords="offset points", ha="right", color=C_INK2, fontsize=9)
        ax.set_ylim(0, max(nominal * 4, float(np.max(dt)) * 1.1))

    # ---- warnings in the footer -----------------------------------------
    notes = list(r.warnings) + burn.quality.problems()
    if notes:
        fig.text(0.045, 0.012, "!  " + "   .   ".join(notes[:4]),
                 color=C_CRIT, fontsize=9, ha="left", va="bottom")

    fig.subplots_adjust(top=0.905, bottom=0.075, left=0.065, right=0.975)
    path = outdir / f"burn_{r.burn_id:03d}_charts.png"
    fig.savefig(path, dpi=160, facecolor=C_SURFACE)
    if show:
        plt.show()
    plt.close(fig)
    return path


def plot_comparison(pairs: list[tuple[Burn, Result]], outdir: Path) -> Path | None:
    """All downloaded burns overlaid on one set of axes."""
    pairs = [(b, r) for b, r in pairs if not r.curve.empty]
    if len(pairs) < 2:
        return None
    fig, ax = plt.subplots(figsize=(11, 6), facecolor=C_SURFACE)
    _style(ax, "Time from fire command [s]", "Thrust above rest [N]", "Burn comparison")
    for i, (b, r) in enumerate(pairs[:8]):
        mass = f"{r.prop_mass_kg * 1000:.0f} g" if r.prop_mass_kg > 0 else "mass n/a"
        isp = f"Isp {r.isp:.0f} s" if _ok(r.isp) else "Isp n/a"
        ax.plot(r.curve["t"], r.curve["thrust"] - r.curve["baseline"],
                color=SERIES[i % len(SERIES)], linewidth=2.0, zorder=4,
                label=f"#{r.burn_id} . {r.motor_designation} . {r.total_impulse:.1f} N.s . "
                      f"{mass} . {isp}")
    leg = ax.legend(loc="upper right", frameon=False, fontsize=9)
    for txt in leg.get_texts():
        txt.set_color(C_INK2)
    fig.tight_layout()
    path = outdir / "burn_comparison.png"
    fig.savefig(path, dpi=160, facecolor=C_SURFACE)
    plt.close(fig)
    return path


# ---------------------------------------------------------------------
#  Data export
# ---------------------------------------------------------------------
COLUMNS = {
    "t": "Time [s]",
    "t_ms": "Time [ms]",
    "thrust": "Thrust [N]",
    "net": "Thrust above rest [N]",
    "raw": "Raw ADC count",
    "pyro": "Pyro energised",
    "cont": "Continuity",
    "sat": "ADC saturated",
}


def _export_frame(burn: Burn, r: Result) -> pd.DataFrame:
    df = burn.df.copy()
    df["net"] = df["thrust"] - r.baseline_n
    out = df[["t", "t_ms", "thrust", "net", "raw", "pyro", "cont", "sat"]].copy()
    for col in ("pyro", "cont", "sat"):
        out[col] = out[col].map({True: "YES", False: "NO"})
    return out.rename(columns=COLUMNS)


def export_csv(burn: Burn, r: Result, outdir: Path) -> Path:
    """CSV that a Czech Excel opens on double-click.

    Semicolon separator, decimal comma, UTF-8 BOM, and a leading `sep=;`
    line that tells Excel how to split the file regardless of what the
    machine's regional settings say.
    """
    out = _export_frame(burn, r)
    # Per-column precision: six digits after the comma is noise on
    # milliseconds but genuinely needed on seconds.
    decimals = {COLUMNS["t"]: 6, COLUMNS["t_ms"]: 3,
                COLUMNS["thrust"]: 4, COLUMNS["net"]: 4}
    for col, dec in decimals.items():
        if col in out.columns:
            out[col] = out[col].map(lambda v, d=dec: f"{v:.{d}f}".replace(".", ","))

    path = outdir / f"burn_{r.burn_id:03d}_data.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write("sep=;\n")
        out.to_csv(fh, sep=";", index=False, lineterminator="\r\n")
    return path


def export_summary_csv(rows: list[tuple[str, str]], r: Result, outdir: Path) -> Path:
    path = outdir / f"burn_{r.burn_id:03d}_summary.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write("sep=;\n")
        fh.write("Quantity;Value\r\n")
        for k, v in rows:
            # decimal point to comma, but only between digits and only for
            # genuinely numeric values (the firmware version is not one)
            if k != "Firmware":
                v = re.sub(r"(?<=\d)\.(?=\d)", ",", v)
            fh.write(f"{k};{v}\r\n")
        # The warnings are the half of the report that says which of the
        # numbers above to trust - they belong in the file people open, not
        # only in the console that has already scrolled past.
        if r.warnings:
            fh.write("\r\n")
            fh.write("WARNINGS;\r\n")
            for note in r.warnings:
                fh.write(f";{note}\r\n")
    return path


def export_xlsx(burn: Burn, r: Result, outdir: Path) -> Path | None:
    """Workbook with a summary, the raw data and a native Excel chart.

    Numbers are stored as numbers, so Excel renders them with whatever
    decimal separator the machine is set to - no manual fixing needed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        return None

    wb = Workbook()

    # --- Summary ------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Static fire #{r.burn_id}"
    ws["A1"].font = Font(size=15, bold=True)
    ws["A2"] = r.motor_designation
    ws["A2"].font = Font(size=12, color="52514E")
    row = 4
    for key, val in summary_rows(burn, r):
        if not key:
            row += 1
            continue
        if key.startswith("-") and key.endswith("-"):
            ws.cell(row=row, column=1, value=key.strip("- ")).font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = PatternFill("solid", fgColor="F0EFEC")
            ws.cell(row=row, column=2).fill = PatternFill("solid", fgColor="F0EFEC")
        else:
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=val).alignment = Alignment(horizontal="right")
        row += 1

    notes = list(r.warnings) + burn.quality.problems()
    if notes:
        row += 1
        ws.cell(row=row, column=1, value="WARNINGS").font = Font(bold=True, color="D03B3B")
        for note in notes:
            row += 1
            ws.cell(row=row, column=1, value="- " + note)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 30

    # --- Data ---------------------------------------------------------
    wsd = wb.create_sheet("Data")
    out = _export_frame(burn, r)
    wsd.append(list(out.columns))
    for cell in wsd[1]:
        cell.font = Font(bold=True)
    for rec in out.itertuples(index=False):
        wsd.append(list(rec))
    for col_idx, name in enumerate(out.columns, start=1):
        letter = get_column_letter(col_idx)
        wsd.column_dimensions[letter].width = max(12, len(name) + 2)
        if "[" in name:
            fmt = "0.000" if ("N]" in name or "s]" in name) else "0.0"
            for cell in wsd[letter][1:]:
                cell.number_format = fmt
    wsd.freeze_panes = "A2"

    # a native Excel chart, so the data stays workable in the spreadsheet
    n = len(out) + 1
    chart = LineChart()
    chart.title = f"Thrust curve - burn {r.burn_id}"
    chart.y_axis.title = "Thrust [N]"
    chart.x_axis.title = "Time [s]"
    chart.height, chart.width = 10, 26
    chart.style = 2
    chart.add_data(Reference(wsd, min_col=3, min_row=1, max_row=n), titles_from_data=True)
    chart.set_categories(Reference(wsd, min_col=1, min_row=2, max_row=n))
    for s in chart.series:
        s.smooth = False
    wsd.add_chart(chart, "K2")

    # --- Rise / decay -------------------------------------------------
    wsp = wb.create_sheet("Rise and decay")
    wsp.append(["Fraction of peak [%]", "Level [N]", "Rise - time [s]", "Decay - time [s]"])
    for cell in wsp[1]:
        cell.font = Font(bold=True)
    for pct in PCT_STEPS:
        rise = r.pct_times.get(pct)
        fall = r.decay_times.get(pct)
        wsp.append([pct, r.peak_thrust * pct / 100.0,
                    rise if _ok(rise) else None,
                    fall if _ok(fall) else None])
    for letter, width, fmt in (("A", 20, "0"), ("B", 14, "0.00"),
                               ("C", 16, "0.0000"), ("D", 16, "0.0000")):
        wsp.column_dimensions[letter].width = width
        for cell in wsp[letter][1:]:
            cell.number_format = fmt

    path = outdir / f"burn_{r.burn_id:03d}_data.xlsx"
    wb.save(path)
    return path


def export_overview_csv(pairs: list[tuple[Burn, Result]], outdir: Path) -> Path:
    """One row per burn - for comparing batches of propellant."""
    recs = []
    for b, r in pairs:
        recs.append({
            "Burn": r.burn_id,
            "Designation": r.motor_designation,
            "Class": r.motor_class,
            "Peak thrust [N]": r.peak_thrust,
            "Average thrust [N]": r.avg_thrust,
            "Total impulse [N.s]": r.total_impulse,
            "Burn time [s]": r.burn_time,
            "Action time [s]": r.action_time,
            "Ignition delay [ms]": r.t_first_motion * 1000.0 if _ok(r.t_first_motion) else None,
            "Rise 10-90 % [ms]": r.rise_10_90 * 1000.0 if _ok(r.rise_10_90) else None,
            "Propellant mass [g]": r.prop_mass_kg * 1000.0,
            "Isp [s]": r.isp if _ok(r.isp) else None,
            "Samples": r.n_samples,
            "Sample rate [Hz]": r.sample_rate_hz,
            "Baseline noise [N]": r.baseline_noise_rms,
        })
    df = pd.DataFrame(recs)
    path = outdir / "all_burns_overview.csv"
    with open(path, "w", encoding="utf-8-sig", newline="") as fh:
        fh.write("sep=;\n")
        df.to_csv(fh, sep=";", decimal=",", index=False,
                  float_format="%.4f", lineterminator="\r\n")
    return path
